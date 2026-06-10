"""
homologar_ppto.py — Incorpora un nuevo proyecto al Control de Áridos.

Uso:
    python homologar_ppto.py CL040
    python homologar_ppto.py CL040 --solo-verificar
    python homologar_ppto.py CL040 --forzar          <- regenera Ppto_APU.xlsx aunque exista

Proceso:
  1. Localiza la carpeta en proyectos/CL{nro}_*/
  2. Detecta el formato del Ppto.xlsx (APU_PPTO directo o Matriz)
  3. Genera Ppto_APU.xlsx si es necesario
  4. Crea proyecto.json si no existe
  5. Copia todo a Data/proyectos/
  6. Verifica la carga con la lógica del dashboard
"""

import sys
import json
import shutil
import re
from pathlib import Path

import openpyxl
from openpyxl import Workbook
import pandas as pd

# ── Rutas base ────────────────────────────────────────────────────────────────
ROOT        = Path(__file__).parent.parent          # Control_Aridos/
PROJS_SRC   = ROOT / "proyectos"                    # Control_Aridos/proyectos/
PROJS_DST   = Path(__file__).parent / "proyectos"   # Control_Aridos/Data/proyectos/

# ── Palabras clave áridos (igual que el dashboard) ────────────────────────────
PALABRAS_ARIDOS = ["arena","grava","bolón","bolon","integral",
                   "base estabilizada","ripio","gravilla","polvo de piedra",
                   "tierra","maicillo"]

def es_arido(t):
    return any(p in str(t).lower() for p in PALABRAS_ARIDOS)

# ── Columnas APU_PPTO (posición exacta que espera el dashboard) ───────────────
COLS_APU = ["CC","Especialidad","Capitulo","Capitulo2","Partida","Destino","Partida2",
            "Codigo","Tipo","Ud","Descripcion","CanPres","PU","ImpPres",
            "Rep1","Rep2","CantPartida","Cantidad","Total",
            "DisponibleMat","CostoMat","MargenMat","NOC",
            "DisponibleSub","CostoSub","MargenSub","NContrato","X1","X2","X3","X4"]


# ══════════════════════════════════════════════════════════════════════════════
# 1. DETECTAR CARPETA DE PROYECTO
# ══════════════════════════════════════════════════════════════════════════════
def encontrar_carpeta(codigo: str) -> Path:
    matches = [d for d in PROJS_SRC.iterdir()
               if d.is_dir() and d.name.upper().startswith(codigo.upper())]
    if not matches:
        raise FileNotFoundError(f"No se encontró carpeta para {codigo} en {PROJS_SRC}")
    return matches[0]


# ══════════════════════════════════════════════════════════════════════════════
# 2. DETECTAR FORMATO DEL PPTO
# ══════════════════════════════════════════════════════════════════════════════
def detectar_formato(ppto_path: Path) -> str:
    """Retorna 'directo', 'matriz', o 'desconocido'."""
    wb = openpyxl.load_workbook(ppto_path, read_only=True, data_only=True)
    hojas = wb.sheetnames
    wb.close()
    if "APU_PPTO" in hojas and "Nombre_Cuenta_Costo" in hojas:
        return "directo"
    if "Matriz" in hojas:
        return "matriz"
    return "desconocido"


# ══════════════════════════════════════════════════════════════════════════════
# 3. TRANSFORMAR FORMATO MATRIZ → APU_PPTO  (detección dinámica de columnas)
# ══════════════════════════════════════════════════════════════════════════════
CC_REGEX = r"^[A-Z]\d+(\.\d+)?$"   # Acepta A03 y también A03.4

def _norm(v):
    return str(v).strip().upper() if v is not None else ""

def _detectar_cabecera(rows):
    """Busca la fila de headers: debe tener NOMBRE PARTIDA + CC/CUENTA COSTO."""
    for i, row in enumerate(rows[:12]):
        r = [_norm(v) for v in row]
        tiene_nombre = "NOMBRE PARTIDA" in r or "NOMBRE_PARTIDA" in r
        tiene_cc     = any(x in r for x in ("CUENTA COSTO", "CC", "C.C.", "CUENTA_COSTO"))
        if tiene_nombre and tiene_cc:
            col_map = {}
            for j, v in enumerate(r):
                if v and v not in col_map:
                    col_map[v] = j
            return i, col_map
    raise ValueError("No se encontro cabecera en Matriz (necesita NOMBRE PARTIDA + CC/CUENTA COSTO)")

def _detectar_col_tipo(col_map, data_rows):
    """Detecta el indice de la columna con valores Material/Maquinaria/Servicios."""
    # Preferencia: E* (inequívoco), luego deteccion por contenido
    if "E*" in col_map:
        return col_map["E*"]
    TIPOS = {"material", "maquinaria", "servicios", "subcontratista", "mano de obra"}
    conteo = {}
    for row in data_rows[:25]:
        if not row or all(x is None for x in row):
            continue
        for j, val in enumerate(row):
            if val and str(val).strip().lower() in TIPOS:
                conteo[j] = conteo.get(j, 0) + 1
    if conteo:
        return max(conteo, key=conteo.get)
    # Fallback por nombre
    for cand in ("IC", "E"):
        if cand in col_map:
            return col_map[cand]
    return None

def _cc_en_excluir(cc_val, excluir_set):
    """True si el CC o su codigo base (A03.4 → A03) esta en la lista de exclusion."""
    if not excluir_set or not cc_val or str(cc_val).strip() in ("", "None"):
        return False
    cc_str = str(cc_val).strip()
    if cc_str in excluir_set:
        return True
    base = cc_str.split(".")[0].strip()
    return base in excluir_set

def _rv(row, idx):
    if idx is None: return None
    return row[idx] if idx < len(row) else None

def _fila_a_apu_din(row, col_map, col_tipo_idx):
    """Mapea una fila de Matriz a 31 columnas APU_PPTO usando indices dinamicos."""
    cc_idx      = col_map.get("CUENTA COSTO", col_map.get("CC", col_map.get("C.C.")))
    desc_idx    = col_map.get("G")
    nombre_idx  = col_map.get("NOMBRE PARTIDA", col_map.get("NOMBRE_PARTIDA"))
    tipo_idx    = col_map.get("TIPO")
    d_idx       = col_map.get("D")
    dstar_idx   = col_map.get("D*")
    f_idx       = col_map.get("F")
    h_idx       = col_map.get("H")
    i_idx       = col_map.get("I")
    j_idx       = col_map.get("J")
    cant_part   = col_map.get("CANTIDAD PARTIDA")
    cant_tot    = col_map.get("CANTIDAD TOTAL PROYECTO")
    rep_idx     = col_map.get("REPETICION", col_map.get("REPETICION"))
    tot_idx     = col_map.get("TOTAL PROYECTO")

    # Buscar 'REPETICION' tolerando tildes
    for k in col_map:
        if k.replace("\xd3","O").replace("\xf3","O").startswith("REPETI"):
            rep_idx = col_map[k]
            break

    tipo_val = _rv(row, col_tipo_idx)
    return [
        _rv(row, cc_idx),       # CC  (se preserva A03.4 tal cual)
        _rv(row, tipo_idx),     # Especialidad
        _rv(row, nombre_idx),   # Capitulo
        _rv(row, nombre_idx),   # Capitulo2
        _rv(row, d_idx),        # Partida
        _rv(row, dstar_idx),    # Destino (D*)
        tipo_val,               # Partida2
        _rv(row, d_idx),        # Codigo
        tipo_val,               # Tipo  ← filtro clave "Material"
        _rv(row, f_idx),        # Ud
        _rv(row, desc_idx),     # Descripcion
        _rv(row, h_idx),        # CanPres
        _rv(row, i_idx),        # PU
        _rv(row, j_idx),        # ImpPres
        _rv(row, cant_part),    # Rep1 (CANTIDAD PARTIDA)
        _rv(row, rep_idx),      # Rep2 (repeticion)
        _rv(row, h_idx),        # CantPartida
        _rv(row, cant_tot),     # Cantidad (CANTIDAD TOTAL PROYECTO)
        _rv(row, tot_idx),      # Total    (TOTAL PROYECTO)
        None, None, None, None,   # DisponibleMat, CostoMat, MargenMat, NOC
        None, None, None, None,   # DisponibleSub, CostoSub, MargenSub, NContrato
        None, None, None, None,   # X1, X2, X3, X4
    ]

def transformar_matriz(ppto_path: Path, salida: Path, nombre_obra: str, excluir_ccs: list = None):
    print(f"  Transformando desde hoja Matriz (deteccion dinamica)...")
    wb_in = openpyxl.load_workbook(ppto_path, read_only=True, data_only=True)
    rows  = list(wb_in["Matriz"].iter_rows(values_only=True))
    wb_in.close()

    hdr_idx, col_map = _detectar_cabecera(rows)
    data_rows        = rows[hdr_idx + 1:]
    col_tipo_idx     = _detectar_col_tipo(col_map, data_rows)

    print(f"  Cabecera en fila {hdr_idx+1} · col Tipo={col_tipo_idx} · {len(col_map)} columnas")

    cc_idx     = col_map.get("CUENTA COSTO", col_map.get("CC", col_map.get("C.C.")))
    tipo_idx   = col_map.get("TIPO")
    nombre_idx = col_map.get("NOMBRE PARTIDA", col_map.get("NOMBRE_PARTIDA"))
    excluir    = set(excluir_ccs or [])

    # Extraer mapa CC → nombre (preservando sub-codigos)
    cc_map       = {}
    current_tipo = ""
    for row in data_rows:
        if not row or all(x is None for x in row):
            continue
        tipo   = str(row[tipo_idx]).strip()   if (tipo_idx   is not None and row[tipo_idx])   else ""
        nombre = str(row[nombre_idx]).strip() if (nombre_idx is not None and row[nombre_idx]) else ""
        cc     = str(row[cc_idx]).strip()     if (cc_idx     is not None and row[cc_idx])     else ""
        if tipo:
            current_tipo = tipo
        # Acepta A03 y A03.4 — preserva la apertura completa
        if cc and cc != "None" and re.match(CC_REGEX, cc) and cc not in cc_map:
            label = f"{cc} - {nombre}" if nombre else f"{cc} - {current_tipo}"
            cc_map[cc] = label

    # Construir libro de salida
    out    = Workbook()
    ws_apu = out.active
    ws_apu.title = "APU_PPTO"
    ws_apu.append([nombre_obra])
    ws_apu.append(["Presupuesto"])
    ws_apu.append(["C.C","Especialidad","Capitulo","Capitulo","Partida","Partida","Partida 2",
                   "Codigo","partida contable","Ud","Resumen","CanPres","Pres","ImpPres",
                   "Repeticiones","Repeticiones","Cant. Partida","Cantidad total","Total Viv",
                   "DISPONIBLE MAT","COSTO MAT","MARGEN MAT","NOC",
                   "DISPONIBLE SUB","COSTO SUB","MARGEN SUB","NContrato",
                   "X1","X2","X3","X4"])
    n = 0
    for row in data_rows:
        if not row or all(x is None for x in row):
            continue
        cc_fila = str(row[cc_idx]).strip() if (cc_idx is not None and row[cc_idx]) else ""
        if _cc_en_excluir(cc_fila, excluir):
            continue
        ws_apu.append(_fila_a_apu_din(row, col_map, col_tipo_idx))
        n += 1

    ws_cc = out.create_sheet("Nombre_Cuenta_Costo")
    ws_cc.append(["Codigo C.C.", "Cuenta de Costo"])
    for cc, nombre in sorted(cc_map.items()):
        if not _cc_en_excluir(cc, excluir):
            ws_cc.append([cc, nombre])

    out.save(salida)
    print(f"  Guardado: {salida.name} ({n} filas · {len(cc_map)} CCs)")
    return len(cc_map)


# ══════════════════════════════════════════════════════════════════════════════
# 4. CREAR proyecto.json SI NO EXISTE
# ══════════════════════════════════════════════════════════════════════════════
def crear_proyecto_json(carpeta: Path, codigo: str, nombre: str, archivo_ppto: str):
    json_path = carpeta / "proyecto.json"
    if json_path.exists():
        print(f"  proyecto.json ya existe — sin cambios")
        return
    data = {"codigo": codigo, "nombre": nombre, "archivo_ppto": archivo_ppto}
    json_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  Creado: proyecto.json")


# ══════════════════════════════════════════════════════════════════════════════
# 5. SINCRONIZAR A Data/proyectos/
# ══════════════════════════════════════════════════════════════════════════════
def sincronizar(carpeta_src: Path, codigo: str):
    carpeta_dst = PROJS_DST / carpeta_src.name
    carpeta_dst.mkdir(parents=True, exist_ok=True)
    archivos = ["proyecto.json", "Descarga_IConstruye.xlsx",
                "Recepcion_Facturación_OC.xlsx"]

    # Agregar el archivo de ppto configurado en proyecto.json
    cfg = json.loads((carpeta_src / "proyecto.json").read_text(encoding="utf-8"))
    archivos.append(cfg["archivo_ppto"])

    copiados = 0
    for nombre in archivos:
        src = carpeta_src / nombre
        dst = carpeta_dst / nombre
        if src.exists():
            shutil.copy2(src, dst)
            copiados += 1
        else:
            print(f"  ADVERTENCIA: {nombre} no encontrado en carpeta fuente")
    print(f"  Sincronizado: {copiados} archivos >> {carpeta_dst}")


# ══════════════════════════════════════════════════════════════════════════════
# 6. VERIFICAR CARGA
# ══════════════════════════════════════════════════════════════════════════════
def verificar(carpeta: Path) -> bool:
    cfg        = json.loads((carpeta / "proyecto.json").read_text(encoding="utf-8"))
    ppto_file  = carpeta / cfg["archivo_ppto"]
    oc_file    = carpeta / "Descarga_IConstruye.xlsx"
    recfac_file= carpeta / "Recepcion_Facturación_OC.xlsx"

    ok = True

    # Ppto
    df_p = pd.read_excel(ppto_file, sheet_name="APU_PPTO",
                         header=None, skiprows=3, names=COLS_APU)
    df_p = df_p.dropna(subset=["Descripcion"])
    df_p = df_p[df_p["Tipo"].astype(str).str.strip() == "Material"]
    df_p = df_p[df_p["Ud"].astype(str).str.upper().str.strip() == "M3"]
    df_p = df_p[df_p["Descripcion"].astype(str).apply(es_arido)]
    df_p["CC"] = df_p["CC"].astype(str).str.strip()
    df_p = df_p[df_p["CC"].str.match(CC_REGEX, na=False)]
    df_p["Cantidad"] = pd.to_numeric(df_p["Cantidad"], errors="coerce").fillna(0)
    df_p["Total"]    = pd.to_numeric(df_p["Total"],    errors="coerce").fillna(0)

    n_ppto = len(df_p)
    tot    = df_p["Total"].sum()
    ccs    = sorted(df_p["CC"].unique())
    print(f"  Ppto áridos  : {n_ppto} ítems · ${tot:,.0f} · CCs: {ccs}")
    if n_ppto == 0:
        print("  ERROR: No se detectaron áridos en el presupuesto")
        ok = False

    # OC
    df_oc = pd.read_excel(oc_file, sheet_name=0, header=None, skiprows=12)
    df_oc.columns = ["N_OC","Nombre_OC","Fecha","Fecha_Despacho","Metodo_Despacho","Obra",
                     "Razon_Social","Rut","Proveedor","Cod_Maestro","Descripcion","Glosa",
                     "CC","Cuenta_Costo","Partida","Unidad","Cantidad","Moneda",
                     "PU","PU_Desc","Descuento","SubTotal","Cant_Recibida","Monto_Recibido",
                     "Devolucion","Saldo_Recibir","Facturado","Monto_Cerrado","Estado"]
    df_oc = df_oc.dropna(subset=["Descripcion"])
    df_oc_a = df_oc[df_oc["Descripcion"].astype(str).apply(es_arido)]
    print(f"  OC áridos    : {len(df_oc_a)} ítems")

    # RecFac
    try:
        df_rf = pd.read_excel(recfac_file, sheet_name="Hoja")
        print(f"  Recep/Factura: {len(df_rf)} filas")
    except Exception as e:
        print(f"  ADVERTENCIA RecFac: {e}")

    # CC names
    df_cc = pd.read_excel(ppto_file, sheet_name="Nombre_Cuenta_Costo")
    df_cc = df_cc.iloc[:, :2].copy(); df_cc.columns = ["CC","Nombre"]
    df_cc = df_cc[df_cc["CC"].astype(str).str.match(CC_REGEX, na=False)]
    print(f"  Cuentas costo: {len(df_cc)} CCs válidos")

    return ok


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    args = sys.argv[1:]
    if not args:
        print(__doc__)
        sys.exit(0)

    codigo        = args[0].upper()
    solo_verificar= "--solo-verificar" in args

    print(f"\n{'='*60}")
    print(f"  Control de Áridos — Incorporando {codigo}")
    print(f"{'='*60}\n")

    # Localizar carpeta
    carpeta = encontrar_carpeta(codigo)
    print(f"Carpeta fuente : {carpeta}\n")

    # Nombre de la obra desde el nombre de carpeta (quita el código del prefijo)
    nombre_obra = re.sub(r"^CL\d+[_\-]?", "", carpeta.name).replace("_", " ")

    ppto_orig = carpeta / "Ppto.xlsx"

    # Leer excluir_ccs desde proyecto.json si existe
    json_path   = carpeta / "proyecto.json"
    excluir_ccs = []
    if json_path.exists():
        cfg_json    = json.loads(json_path.read_text(encoding="utf-8"))
        excluir_ccs = cfg_json.get("excluir_ccs", [])
    if excluir_ccs:
        print(f"CCs excluidas  : {excluir_ccs}\n")

    if solo_verificar:
        print("--- Verificando Data/proyectos/ ---")
        carpeta_dst = PROJS_DST / carpeta.name
        verificar(carpeta_dst)
        return

    # Determinar formato y preparar Ppto_APU.xlsx
    if ppto_orig.exists():
        fmt = detectar_formato(ppto_orig)
        print(f"Formato Ppto   : {fmt}")

        if fmt == "directo":
            print("  Ppto.xlsx tiene APU_PPTO — se usa directamente")
            ppto_apu = ppto_orig
            nombre_ppto_cfg = "Ppto.xlsx"
        elif fmt == "matriz":
            ppto_apu = carpeta / "Ppto_APU.xlsx"
            forzar   = "--forzar" in args
            if ppto_apu.exists() and not forzar:
                print("  Ppto_APU.xlsx ya existe — se preserva (usar --forzar para regenerar)")
            else:
                print("Generando Ppto_APU.xlsx desde Matriz...")
                transformar_matriz(ppto_orig, ppto_apu, nombre_obra, excluir_ccs)
            nombre_ppto_cfg = "Ppto_APU.xlsx"
        else:
            print("  ADVERTENCIA: formato no reconocido — se asume APU_PPTO en Ppto.xlsx")
            ppto_apu = ppto_orig
            nombre_ppto_cfg = "Ppto.xlsx"
    else:
        ppto_apu = carpeta / "Ppto_APU.xlsx"
        if not ppto_apu.exists():
            print("ERROR: No se encontró Ppto.xlsx ni Ppto_APU.xlsx en la carpeta")
            sys.exit(1)
        nombre_ppto_cfg = "Ppto_APU.xlsx"

    # Crear proyecto.json si no existe
    print("\nConfigurando proyecto.json...")
    crear_proyecto_json(carpeta, codigo, nombre_obra, nombre_ppto_cfg)

    # Sincronizar a Data/proyectos/
    print("\nSincronizando a Data/proyectos/...")
    sincronizar(carpeta, codigo)

    # Verificar
    print("\n--- Verificación final ---")
    carpeta_dst = PROJS_DST / carpeta.name
    ok = verificar(carpeta_dst)

    print()
    if ok:
        print(f"[OK] {codigo} incorporado correctamente")
        print(f"   Levantar dashboard: py -m streamlit run app.py --server.port 8505")
    else:
        print(f"[ERROR] Hay errores -- revisar los mensajes anteriores")
    print()


if __name__ == "__main__":
    main()
